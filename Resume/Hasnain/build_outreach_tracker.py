import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

hf = Font(bold=True, color='FFFFFF', size=11)
hfill = PatternFill('solid', fgColor='2F5496')
dfill = PatternFill('solid', fgColor='1B3A5C')
dfont = Font(bold=True, color='FFFFFF', size=12)
lgray = PatternFill('solid', fgColor='F5F5F5')
gfill = PatternFill('solid', fgColor='C6EFCE')
yfill = PatternFill('solid', fgColor='FFF2CC')
pfill = PatternFill('solid', fgColor='548235')
rfill = PatternFill('solid', fgColor='BF4B28')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
bb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('medium'))
wrap = Alignment(wrap_text=True, vertical='top')

def style_hdr(ws, r, mc, fill=None):
    for c in range(1, mc+1):
        cl = ws.cell(r, c); cl.font = hf; cl.fill = fill or hfill
        cl.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center'); cl.border = tb

def style_row(ws, r, mc, f=None):
    for c in range(1, mc+1):
        cl = ws.cell(r, c); cl.border = tb; cl.alignment = wrap
        if f: cl.fill = f

start = datetime(2026, 8, 8)

# =========================================================
# SHEET 1: HEALTHCARE AI — 8 companies/day x 10 days = 80
# Real African & global healthtech companies/startups
# =========================================================

# (company, city, what_they_do, pain_point, pitch, who_to_find, size, stage)
companies = [
    # DAY 1 — Nigerian Healthtech
    ('Helium Health', 'Lagos, Nigeria',
     'Healthcare SaaS. EMR (electronic medical records), hospital management, telemedicine across 10 African countries',
     'Hospitals still on paper. Doctors can\'t search patient history fast. No AI triage. Appointment booking manual. Staff training expensive',
     'AI patient triage chatbot + Medical records RAG search + AI appointment scheduler + AI staff training bot',
     'CTO, Co-Founder, VP Engineering, Head of Product', '~100', 'Series B'),
    ('Lifestores Healthcare', 'Lagos, Nigeria',
     'Pharmacy distribution platform. Connects pharmacies to drug suppliers. 1000+ pharmacies',
     'Pharmacies ask same questions: drug availability, pricing, alternatives. Order tracking manual. No demand prediction',
     'AI drug availability chatbot + AI demand forecasting + AI alternative drug recommender + Order tracking bot',
     'CTO, Co-Founder, Head of Engineering', '~30', 'Series A'),
    ('Remedial Health', 'Lagos, Nigeria',
     'Pharmacy supply chain. B2B marketplace for pharmaceutical products across Nigeria',
     'Drug procurement manual. Counterfeit drug risk. Pharmacies can\'t predict demand. No smart ordering',
     'AI drug verification system + AI demand prediction + AI smart ordering chatbot + Supply chain analytics',
     'CTO, Co-Founder, Head of Engineering', '~25', 'Series A'),
    ('54gene', 'Lagos, Nigeria',
     'Genomics & precision medicine for Africa. African genetic data for drug development',
     'Massive genomic data but no AI analysis tools. Researchers need AI to find patterns. Data scattered',
     'AI genomic data analysis dashboard + RAG on research papers + AI research assistant chatbot',
     'CTO, Head of AI/Data, Head of Engineering', '~50', 'Series B'),
    ('Reliance HMO', 'Lagos, Nigeria',
     'Health insurance & HMO platform. Digital health insurance for individuals and companies',
     'Claims processing manual. Members can\'t check benefits easily. Provider network search is basic',
     'AI claims processing bot + AI benefits checker chatbot + AI provider matching + Smart claims routing',
     'CTO, Co-Founder, Head of Engineering', '~40', 'Series A'),
    ('MDaaS Global', 'Lagos, Nigeria',
     'Diagnostic centers. Affordable medical diagnostics (labs, scans) for underserved communities',
     'Patients don\'t know which test they need. Results interpretation confusing. Appointment booking chaotic',
     'AI symptom-to-test recommender + AI results explainer chatbot + AI appointment scheduler',
     'CTO, Co-Founder, Head of Technology', '~30', 'Series A'),
    ('Kangpe / Doctall', 'Lagos, Nigeria',
     'Telemedicine platform. Online doctor consultations, lab tests, pharmacy delivery',
     'Patients wait in queue for teleconsult. Triage non-existent. Follow-up manual. Prescription questions unanswered',
     'AI patient triage (pre-consultation) + AI follow-up chatbot + AI prescription FAQ bot + Smart doctor matching',
     'CTO, Co-Founder, Head of Engineering', '~25', 'Series A'),
    ('WellaHealth', 'Lagos, Nigeria',
     'Micro health insurance. Affordable malaria & health coverage via USSD/mobile for mass market',
     'Policy questions from low-tech users. Claims via phone calls. No self-service. Agent support overwhelmed',
     'AI USSD/WhatsApp claims bot + AI policy explainer (simple language) + AI agent assist tool',
     'CTO, Co-Founder, Lead Engineer', '~15', 'Seed'),

    # DAY 2 — Egyptian Healthtech
    ('Vezeeta', 'Cairo, Egypt',
     'Doctor booking platform. Largest in MENA. Telemedicine, prescriptions, insurance. 5M+ users',
     'Millions of patients: "which doctor for my problem?" "how to cancel?" "insurance coverage?" All manual support',
     'AI doctor matching chatbot (symptom-based) + AI appointment manager + AI insurance checker + Patient RAG on health FAQs',
     'CTO, VP Engineering, Head of Product, Head of AI', '~300', 'Series D'),
    ('Yodawy', 'Cairo, Egypt',
     'Digital pharmacy. Online medicine ordering, prescription management, home delivery',
     'Drug availability questions. Prescription confusion. Delivery tracking. Drug interaction queries. All manual',
     'AI pharmacy chatbot + AI drug interaction checker + AI prescription assistant + Delivery tracking bot',
     'CTO, Head of Engineering, Head of Product', '~60', 'Series B'),
    ('Chefaa', 'Cairo, Egypt',
     'Pharmacy delivery app. Medicine ordering & chronic disease management',
     'Chronic patients need refill reminders. Drug interaction worries. "Is this available?" queries pile up',
     'AI refill reminder system + AI drug interaction RAG + AI availability chatbot + Chronic care assistant',
     'CTO, Co-Founder, Lead Engineer', '~20', 'Series A'),
    ('Almouneer', 'Cairo, Egypt',
     'Diabetes management platform. Helping diabetic patients track & manage their condition',
     'Patients need daily guidance. Diet questions. Medication timing. Blood sugar interpretation. Doctors can\'t answer everyone',
     'AI diabetes management chatbot + AI diet advisor + AI blood sugar interpreter + Medical RAG on diabetes knowledge',
     'CTO, Co-Founder, Head of Product', '~15', 'Seed'),
    ('DabaDoc', 'Casablanca, Morocco',
     'Doctor booking & telemedicine for Morocco & North Africa. 6000+ doctors',
     'Patients searching for right doctor. Appointment management. Medical queries unanswered. French/Arabic support needed',
     'AI doctor matching (multilingual) + AI appointment chatbot + AI symptom checker + French/Arabic NLP',
     'CTO, Co-Founder, Head of Engineering', '~25', 'Series A'),
    ('Nawah Scientific', 'Cairo, Egypt',
     'Clinical research & diagnostics. Lab services, clinical trials support',
     'Researchers sifting through papers manually. Lab result interpretation time-consuming. No AI research tools',
     'AI research paper RAG + AI lab result interpreter + AI clinical trial matcher + Research assistant chatbot',
     'CTO, Co-Founder, Head of Data', '~20', 'Seed'),
    ('Bypa-ss', 'Cairo, Egypt',
     'Cardiac surgery platform. Connects heart patients with cardiac surgeons across borders',
     'Patients confused about surgery options. Cost queries. Doctor selection. Post-surgery follow-up. All via phone/email',
     'AI patient inquiry chatbot + AI surgeon matching + AI cost estimator + Post-surgery follow-up bot + Medical RAG',
     'CTO, Co-Founder, Lead Engineer', '~15', 'Seed'),
    ('Ezaby Pharmacy', 'Cairo, Egypt',
     'Largest pharmacy chain in Egypt going digital. Online ordering, delivery',
     'Thousands of customer queries daily. Drug availability. Delivery status. Drug alternatives. Pharmacist time wasted on FAQs',
     'AI pharmacy FAQ chatbot + AI drug alternative recommender + AI delivery tracker + Smart escalation to pharmacist',
     'Head of Digital, Head of Technology, CTO', '~500', 'Enterprise'),

    # DAY 3 — Kenyan Healthtech
    ('M-TIBA', 'Nairobi, Kenya',
     'Mobile health wallet by Safaricom + PharmAccess. Health savings & insurance via M-Pesa. 5M+ users',
     'Users don\'t understand health wallet. Claim process confusing. Benefits unclear. Support via call center = expensive',
     'AI health wallet explainer chatbot + AI claims guide bot + AI benefits checker + WhatsApp/USSD support',
     'Head of Product, Head of Technology, CTO', '~30', 'Corporate'),
    ('Ilara Health', 'Nairobi, Kenya',
     'Diagnostic devices for clinics. AI-powered diagnostics hardware + software for African clinics',
     'Clinic staff need training on devices. Result interpretation support. Device troubleshooting. All done by field team',
     'AI training chatbot (RAG on device manuals) + AI result interpretation assistant + AI troubleshooting bot',
     'CTO, Co-Founder, Head of Engineering', '~30', 'Series A'),
    ('Mydawa', 'Nairobi, Kenya',
     'Online pharmacy. Licensed e-pharmacy in Kenya. Prescription delivery, health products',
     'Prescription questions. Drug availability. Delivery tracking. Drug interaction queries. Small support team',
     'AI prescription assistant chatbot + AI drug checker + AI delivery tracking bot + Pharmacy knowledge RAG',
     'CTO, Co-Founder, Head of Engineering', '~25', 'Series A'),
    ('Ponea Health', 'Nairobi, Kenya',
     'Healthcare marketplace. Connects patients to affordable health services & products',
     'Patients searching for right service/product. Price comparison confusing. Provider selection. No guidance',
     'AI health service recommender + AI price comparison chatbot + AI provider matching + Symptom-to-service guide',
     'CTO, Co-Founder, Lead Engineer', '~20', 'Seed'),
    ('Access Afya', 'Nairobi, Kenya',
     'Affordable primary care clinics for low-income communities in Kenya',
     'High patient volume. Triage by nurses overwhelmed. Patient education needed. Follow-up reminders manual',
     'AI patient triage assistant + AI health education chatbot + AI follow-up reminder system + Clinical decision RAG',
     'CTO, Head of Technology, Clinic Director', '~30', 'Series A'),
    ('Savannah Informatics', 'Nairobi, Kenya',
     'Health information exchange. Connecting hospitals, labs, pharmacies digitally',
     'Data integration complex. Hospital staff need help using system. Interoperability issues. Training expensive',
     'AI onboarding/training chatbot (RAG on system docs) + AI data query assistant + AI interoperability helper',
     'CTO, Co-Founder, Head of Engineering', '~20', 'Series A'),
    ('Afya Rekod', 'Nairobi, Kenya',
     'Blockchain-based health records. Patient-owned medical records',
     'Patients don\'t understand their records. Sharing records confusing. Health history scattered. No AI layer',
     'AI medical record explainer chatbot + AI health history summarizer + AI record sharing assistant',
     'CTO, Co-Founder, Lead Engineer', '~10', 'Seed'),
    ('Jacaranda Health', 'Nairobi, Kenya',
     'Maternal healthcare. Affordable childbirth, pregnancy support, maternal clinics',
     'Pregnant women have 100s of questions. Follow-up appointments missed. Nutrition guidance needed. Nurses overwhelmed',
     'AI pregnancy support chatbot + AI appointment reminder + AI nutrition advisor + Maternal health RAG',
     'CTO, Head of Technology, Clinical Director', '~30', 'Series A'),

    # DAY 4 — South African Healthtech
    ('Envisionit Deep AI', 'Johannesburg, South Africa',
     'AI radiology. AI-powered X-ray and medical image analysis for TB, pneumonia detection',
     'Radiologists overwhelmed. Results take days. Rural clinics have no radiologists. Need faster AI screening',
     'Improve AI diagnostic models + AI result explanation chatbot + AI clinical workflow integration + RAG on radiology guidelines',
     'CTO, Co-Founder, Head of AI', '~20', 'Series A'),
    ('Vula Mobile', 'Cape Town, South Africa',
     'Clinical referral app. Connects primary care doctors to specialists for consultations',
     'Referral process slow. Doctors need guidance on when to refer. Follow-up lost. Communication via phone calls',
     'AI referral recommendation system + AI clinical decision support + AI follow-up tracker + Specialist matching',
     'CTO, Co-Founder, Head of Engineering', '~15', 'Series A'),
    ('Wysa', 'Cape Town / Global',
     'AI mental health chatbot. CBT-based therapy, mood tracking, anxiety support',
     'Growing mental health crisis in Africa. Existing chatbot needs multi-agent upgrade. Personalization limited',
     'Multi-agent therapy system (your 3-agent expertise) + RAG on therapy protocols + Personalized treatment paths',
     'CTO, Head of AI, Head of Engineering', '~50', 'Series B'),
    ('hearX Group', 'Pretoria, South Africa',
     'Digital hearing health. Smartphone-based hearing tests, teleaudiology',
     'Hearing test results confusing for patients. Follow-up manual. Specialist referral process slow. Patient education needed',
     'AI hearing result explainer chatbot + AI specialist matching + AI follow-up system + Patient education RAG',
     'CTO, Co-Founder, Head of Product', '~30', 'Series A'),
    ('Quro Medical', 'Cape Town, South Africa',
     'Hospital-at-home platform. Remote patient monitoring, virtual wards',
     'Patients at home need constant check-ins. Nurses monitoring multiple patients. Alert fatigue. Escalation manual',
     'AI patient monitoring chatbot + AI alert triage system + AI nurse assist dashboard + Smart escalation',
     'CTO, Co-Founder, Head of Engineering', '~20', 'Series A'),
    ('Vula / RecoMed', 'Cape Town, South Africa',
     'Doctor booking & practice management for South African healthcare providers',
     'Patients booking appointments. No-show prediction. Patient reminders. Insurance verification. All manual',
     'AI appointment chatbot + AI no-show predictor + AI insurance verification + Patient reminder system',
     'CTO, Co-Founder, Head of Product', '~20', 'Series A'),
    ('Kena Health', 'Johannesburg, South Africa',
     'Telehealth & online doctor consultations for South Africa',
     'Patients waiting for teleconsult. No pre-triage. Repeat prescription questions. Follow-up lost. Growing user base',
     'AI pre-triage chatbot + AI prescription FAQ bot + AI follow-up scheduler + Smart doctor queue management',
     'CTO, Co-Founder, Lead Engineer', '~15', 'Seed'),
    ('Impilo', 'Johannesburg, South Africa',
     'Chronic disease management platform. Diabetes, hypertension, HIV management',
     'Chronic patients need daily guidance. Medication adherence low. Diet questions constant. Doctors can\'t follow up everyone',
     'AI chronic care chatbot + AI medication reminder + AI diet advisor + Health tracking RAG + Adherence predictor',
     'CTO, Co-Founder, Head of Product', '~15', 'Seed'),

    # DAY 5 — Ghanaian + Rwandan + Other Healthtech
    ('mPharma', 'Accra, Ghana',
     'Pharmacy management & drug supply chain. 350+ pharmacies across 9 African countries',
     'Pharmacies: drug stock queries, ordering questions, pricing confusion. Training on system. Support spread across 9 countries',
     'AI pharmacy support chatbot + AI drug stock checker + AI pricing assistant + AI training bot (RAG on manuals)',
     'CTO, Head of Engineering, Head of Data', '~250', 'Series D'),
    ('Zipline', 'Accra, Ghana / Kigali, Rwanda',
     'Drone delivery of medical supplies. Blood, vaccines, medicines delivered by drone to remote areas',
     'Hospital staff ordering supplies. Delivery tracking. Inventory management at hospitals. Demand prediction',
     'AI medical supply ordering chatbot + AI delivery tracking + AI hospital inventory predictor + Demand forecasting',
     'CTO, Head of AI, Head of Engineering', '~500', 'Series F'),
    ('Redbird Health Tech', 'Accra, Ghana',
     'Point-of-care diagnostics. Rapid diagnostic devices for community pharmacies & clinics',
     'Pharmacy staff need help interpreting results. Device training needed. Quality control. Patient counseling',
     'AI result interpretation assistant + AI training chatbot (RAG on protocols) + AI patient counseling bot',
     'CTO, Co-Founder, Head of Engineering', '~20', 'Series A'),
    ('eHealth Africa', 'Kano, Nigeria / Accra',
     'Digital health solutions for governments & NGOs. Disease surveillance, immunization tracking',
     'Health workers in field need real-time guidance. Data entry errors. Disease outbreak detection slow. Training gaps',
     'AI field worker chatbot + AI disease surveillance alerts + AI data quality checker + Training RAG',
     'CTO, Head of Engineering, Head of AI', '~100', 'NGO/Grant'),
    ('Babyl Rwanda (Babylon Health)', 'Kigali, Rwanda',
     'AI triage & telemedicine for Rwanda. Government partnership. 2M+ registered users',
     'Massive user base. AI triage needs improvement. Multi-language (Kinyarwanda + English). Follow-up management',
     'Improved multi-agent triage system + RAG on Rwandan health guidelines + AI follow-up scheduler + Multi-language NLP',
     'CTO, Head of AI, Head of Engineering', '~40', 'Series B'),
    ('Viebeg Technologies', 'Kigali, Rwanda',
     'Medical supply chain platform. Connects hospitals to medical equipment suppliers',
     'Hospitals ordering equipment. Stock availability queries. Delivery tracking. Equipment training needed',
     'AI supply ordering chatbot + AI stock availability checker + AI delivery tracker + Equipment training RAG',
     'CTO, Co-Founder, Lead Engineer', '~15', 'Seed'),
    ('PharmAccess', 'Pan-Africa (Kenya, Nigeria, Ghana, Tanzania)',
     'Health financing & digital health programs. Mobile health insurance, quality improvement',
     'Health program beneficiaries confused about benefits. Claims process unclear. Quality reporting manual',
     'AI beneficiary support chatbot + AI claims guide + AI quality reporting assistant + Program knowledge RAG',
     'Head of Digital, CTO, Head of Programs', '~100', 'NGO'),
    ('Flatiron Health Africa (partnership)', 'Pan-Africa',
     'Oncology data platform. Cancer research & treatment data across African hospitals',
     'Cancer data scattered. Research teams need AI tools. Clinical trial matching for patients. Data quality issues',
     'AI clinical trial matcher chatbot + AI research data RAG + AI data quality assistant + Cancer knowledge base',
     'Head of Engineering, Head of Data, CTO', '~30', 'Corporate'),

    # DAY 6 — Telemedicine & Mental Health
    ('Africa Health Holdings', 'Pan-Africa',
     'Hospital network across Africa. Clinics & hospitals in multiple countries going digital',
     'Each hospital has different systems. Staff training across sites. Patient records not searchable. No AI',
     'AI patient records RAG search + AI staff training bot + AI clinical decision support + Cross-hospital analytics',
     'CTO, Head of Digital, VP Technology', '~200', 'Series B'),
    ('HealthPlus', 'Lagos, Nigeria',
     'Pharmacy retail chain. One of Nigeria\'s largest pharmacy chains going digital',
     'In-store pharmacists overwhelmed with questions. Drug interaction queries. Product recommendations. Online ordering growing',
     'AI pharmacist assistant chatbot + AI drug interaction RAG + AI product recommender + Online order support bot',
     'Head of Digital, CTO, Head of Operations', '~300', 'Enterprise'),
    ('MyDawa (expanded)', 'Nairobi, Kenya',
     'Online pharmacy expanding into chronic disease management & subscription medicines',
     'Chronic patients need medication guidance. Refill management. Side effect questions. Drug interaction worries',
     'AI chronic medication chatbot + AI side effect checker (medical RAG) + AI refill scheduler + Drug interaction bot',
     'CTO, Head of Product, Lead Engineer', '~30', 'Series A'),
    ('Zoie Health', 'Cape Town, South Africa',
     'Women\'s health platform. Telehealth, wellness, hormone health for women in Africa',
     'Women\'s health questions sensitive. Need private AI support. Symptom tracking. Doctor matching. Education gaps',
     'AI women\'s health chatbot (sensitive, private) + AI symptom tracker + AI doctor matching + Health education RAG',
     'CTO, Co-Founder, Head of Product', '~15', 'Seed'),
    ('Penda Health', 'Nairobi, Kenya',
     'Affordable outpatient clinics. 20+ clinics in Nairobi for everyday health needs',
     'High patient volume. Triage by receptionists. Same symptoms questions daily. Lab result follow-up manual',
     'AI triage chatbot (pre-visit) + AI lab result explainer + AI follow-up scheduler + Clinical knowledge RAG',
     'CTO, Head of Technology, Medical Director', '~100', 'Series B'),
    ('mDoc', 'Lagos, Nigeria',
     'Digital health coaching. Chronic disease management via health coaches & technology',
     'Health coaches overloaded. Patients need 24/7 support. Diet/exercise guidance repeated. Medication adherence tracking',
     'AI health coaching chatbot + AI diet & exercise advisor + AI medication reminder + Chronic disease RAG',
     'CTO, Co-Founder, Head of Product', '~20', 'Seed'),
    ('TeleDr', 'Lagos, Nigeria',
     'Telemedicine startup. Virtual doctor consultations for Nigerians',
     'Pre-consultation triage missing. Patients describing symptoms inefficiently. Follow-up lost. Prescription queries',
     'AI symptom collector chatbot (pre-consult) + AI follow-up bot + AI prescription FAQ + Smart doctor routing',
     'CTO, Co-Founder, Lead Engineer', '~10', 'Seed'),
    ('Cardo Health', 'Cairo, Egypt',
     'Cardiovascular health platform. Heart health monitoring & management',
     'Heart patients need constant monitoring guidance. Medication questions. Emergency escalation. Lifestyle advice',
     'AI cardiac care chatbot + AI medication guide + AI emergency escalation system + Heart health RAG',
     'CTO, Co-Founder, Head of Engineering', '~15', 'Seed'),

    # DAY 7 — Pharma, Drug Discovery, Clinical Trials
    ('DrugStoc', 'Lagos, Nigeria',
     'Pharmaceutical distribution. B2B drug marketplace connecting manufacturers to pharmacies/hospitals',
     'Pharmacies ordering drugs manually. Price comparison difficult. Stock availability unknown. Delivery tracking manual',
     'AI drug ordering chatbot + AI price comparison tool + AI stock predictor + AI delivery tracker + Smart reordering',
     'CTO, Co-Founder, Head of Engineering', '~25', 'Series A'),
    ('Field Intelligence (Shelf Life)', 'Lagos, Nigeria',
     'Pharmacy distribution & analytics. Data-driven drug supply chain for Africa',
     'Drug demand prediction manual. Pharmacy analytics missing. Stock-outs common. Expiry waste high',
     'AI drug demand predictor + AI pharmacy analytics dashboard + AI stock-out preventer + AI expiry tracker',
     'CTO, Co-Founder, Head of Data', '~30', 'Series A'),
    ('RxAll', 'Lagos, Nigeria',
     'Drug authentication. AI-powered device that detects counterfeit medicines',
     'Counterfeit drug detection needs scale. Consumer education about fake drugs. Reporting system. Data analytics',
     'AI drug verification chatbot + AI consumer education bot + AI counterfeit reporting system + Analytics dashboard',
     'CTO, Co-Founder, Head of AI', '~15', 'Seed'),
    ('Triomics', 'Cairo, Egypt',
     'Cancer genomics & clinical trial matching using AI',
     'Cancer patients can\'t find right clinical trials. Genomic data analysis complex. Doctor-patient matching for trials',
     'AI clinical trial matcher chatbot + AI genomic data RAG + AI patient-trial matching + Research assistant',
     'CTO, Co-Founder, Head of AI', '~15', 'Seed'),
    ('Injini', 'Cape Town, South Africa',
     'Edtech + healthtech accelerator. Portfolio of health-education startups',
     'Portfolio companies all need AI features. Accelerator wants to offer AI services. Each company has different needs',
     'AI consulting for portfolio: chatbots, RAG, triage systems. Offer bulk deal for multiple startups',
     'Managing Director, Head of Programs, CTO', '~15', 'Accelerator'),
    ('SASdoc', 'Johannesburg, South Africa',
     'Doctor-patient communication platform. Secure messaging, file sharing, consultations',
     'Doctors managing multiple patient conversations. No AI prioritization. Same medical questions repeated',
     'AI message prioritization + AI common questions auto-responder + AI clinical summary generator + Medical RAG',
     'CTO, Co-Founder, Lead Engineer', '~10', 'Seed'),
    ('Waspito', 'Douala, Cameroon',
     'Telemedicine for Francophone Africa. Online doctor consultations in French',
     'French-speaking patients need AI triage. Doctor queue management. Follow-up in French. No AI support',
     'AI French triage chatbot + AI doctor matching + AI follow-up bot + French medical RAG',
     'CTO, Co-Founder, Head of Engineering', '~15', 'Seed'),
    ('Healthlane', 'Lagos, Nigeria',
     'Health benefits management. Employee health benefits platform for companies',
     'Employees don\'t understand benefits. HR teams answering same questions. Claims process confusing. No self-service',
     'AI benefits explainer chatbot + AI claims guide + AI provider finder + HR health analytics dashboard',
     'CTO, Co-Founder, Head of Product', '~15', 'Seed'),

    # DAY 8 — Medical Records, EMR, Hospital Management
    ('Sema (Helium Health competitor)', 'Lagos, Nigeria',
     'Hospital management system. EMR, billing, inventory for Nigerian hospitals',
     'Hospital staff not tech-savvy. EMR training expensive. Billing errors. Inventory management manual',
     'AI EMR training chatbot (RAG on user guide) + AI billing assistant + AI inventory predictor + Staff onboarding bot',
     'CTO, Co-Founder, Head of Engineering', '~20', 'Seed'),
    ('Kira (by Zipline spin)', 'Kigali, Rwanda',
     'Health logistics platform for Rwanda. Supply chain management for medical supplies',
     'Hospital ordering complex. Supply matching to demand. Expiry management. Emergency restocking',
     'AI supply ordering chatbot + AI demand-supply matcher + AI expiry tracker + Emergency restock alerts',
     'CTO, Head of Engineering, Lead Engineer', '~15', 'Seed'),
    ('Lifebank', 'Lagos, Nigeria',
     'Emergency medical logistics. Blood, oxygen, medical supplies delivery to hospitals in emergencies',
     'Emergency ordering chaotic. Blood type matching. Delivery routing under pressure. Hospital inventory unknown',
     'AI emergency ordering chatbot + AI blood type matcher + AI priority routing + Hospital inventory predictor',
     'CTO, Co-Founder, Head of Engineering', '~30', 'Series A'),
    ('TraceRx', 'Lagos, Nigeria',
     'Pharmaceutical supply chain tracking. Blockchain + AI for drug traceability',
     'Drug supply chain opaque. Counterfeit infiltration. Tracking manual. Compliance reporting time-consuming',
     'AI supply chain analytics + AI counterfeit risk scoring + AI compliance reporter + Drug tracking chatbot',
     'CTO, Co-Founder, Head of AI', '~15', 'Seed'),
    ('Aajoh', 'Nairobi, Kenya',
     'Healthcare data analytics. Population health insights for African healthcare providers',
     'Health data scattered across systems. No AI insights. Manual reporting. Population health trends not visible',
     'AI health data analytics dashboard + AI trend predictor + AI report generator + Data query RAG chatbot',
     'CTO, Co-Founder, Head of Data', '~15', 'Seed'),
    ('Ingress Healthcare', 'Cape Town, South Africa',
     'Healthcare IT solutions. Integration, data management, clinical systems for SA hospitals',
     'Hospital system integration complex. Staff training on multiple systems. Data migration projects. Support tickets',
     'AI integration support chatbot + AI training bot (RAG on system docs) + AI data migration assistant',
     'CTO, Head of Engineering, Solution Architect', '~25', 'Series A'),
    ('CareAfya', 'Nairobi, Kenya',
     'Affordable primary healthcare clinics. Community health services in Kenya',
     'High patient volume. Community health workers need decision support. Patient follow-up manual. Health education needed',
     'AI community health worker chatbot + AI clinical decision support + AI patient follow-up + Health education RAG',
     'CTO, Head of Technology, Medical Director', '~25', 'Series A'),
    ('Ubenwa', 'Lagos, Nigeria / Montreal',
     'AI for newborn health. Cry analysis for early detection of birth asphyxia',
     'Newborn health monitoring in low-resource settings. Need AI diagnostic tools. Parent education. Clinical workflow integration',
     'AI parent education chatbot + AI clinical workflow integration + AI monitoring alerts + Neonatal health RAG',
     'CTO, Co-Founder, Head of AI', '~15', 'Seed'),

    # DAY 9 — Health Insurance, HMO, Benefits
    ('Hygeia HMO', 'Lagos, Nigeria',
     'Health insurance / HMO. One of Nigeria\'s largest HMOs. Corporate & individual health plans',
     'Members calling for provider lists, benefits, claims. Call center overwhelmed. Same FAQs 1000x/day',
     'AI member support chatbot + AI provider finder + AI claims status bot + AI benefits explainer + RAG on policies',
     'CTO, Head of Digital, Head of Operations', '~300', 'Enterprise'),
    ('AXA Mansard Health', 'Lagos, Nigeria',
     'Health insurance by AXA in Nigeria. Corporate health plans, individual coverage',
     'Policy holders confused about coverage. Claims process complex. Provider network search basic. Pre-authorization manual',
     'AI coverage checker chatbot + AI claims guide + AI provider matcher + AI pre-authorization assistant',
     'Head of Digital, CTO, Head of Claims', '~500', 'Corporate'),
    ('Cassava Micro Insurance', 'Lagos, Nigeria',
     'Micro health insurance for mass market via mobile. Partnership with MTN',
     'Low-tech users. Policy understanding near zero. Claims via USSD/phone. Agent support expensive',
     'AI USSD/WhatsApp insurance chatbot (simple language) + AI claims bot + AI policy explainer + Agent assist',
     'CTO, Head of Product, Head of Technology', '~30', 'Series A'),
    ('Resolution Insurance', 'Nairobi, Kenya',
     'Health & general insurance for East Africa. Corporate & individual plans',
     'Claims processing slow. Members asking about coverage. Provider search. Pre-authorization bottleneck',
     'AI claims automation + AI member chatbot + AI provider finder + AI pre-auth assistant + Policy RAG',
     'CTO, Head of Digital, Head of Claims', '~200', 'Enterprise'),
    ('Discovery Health (digital)', 'Johannesburg, South Africa',
     'SA\'s largest health insurer going more digital. Vitality wellness program',
     'Millions of members. Claims queries. Vitality points questions. Provider search. Wellness guidance needed',
     'AI member support chatbot + AI claims tracker + AI wellness advisor + AI provider matcher + Vitality RAG',
     'Head of Digital, Head of AI, VP Technology', '~5000', 'Public'),
    ('MedSaf', 'Lagos, Nigeria',
     'Medication safety platform. Helps hospitals avoid medication errors',
     'Medication errors kill. Nurses need drug dosage guidance. Drug interaction checking slow. Alert fatigue',
     'AI drug dosage calculator + AI interaction checker (medical RAG) + AI alert prioritization + Nurse assist chatbot',
     'CTO, Co-Founder, Head of Product', '~15', 'Seed'),
    ('Lami', 'Nairobi, Kenya',
     'Insurance API. Embedded insurance for African platforms. Any app can offer insurance',
     'Partners integrating insurance API. Support tickets for integration. Onboarding questions. Same docs questions',
     'AI developer support chatbot (RAG on API docs) + AI onboarding wizard + AI integration troubleshooter',
     'CTO, Co-Founder, Head of Engineering', '~20', 'Series A'),
    ('Turaco', 'Nairobi, Kenya',
     'Embedded insurance. APIs for companies to offer insurance to their customers',
     'Client companies need integration support. Claims processing for embedded products. Customer support questions',
     'AI integration support chatbot + AI claims processor + AI customer FAQ bot + Partner analytics dashboard',
     'CTO, Co-Founder, Lead Engineer', '~20', 'Seed'),

    # DAY 10 — Global Health + NGOs serving Africa
    ('Andela Health (tech talent for health)', 'Pan-Africa',
     'Tech talent platform. Health companies hiring African developers for health tech projects',
     'Health companies need AI developers. Matching health domain + AI skills. Screening health tech knowledge',
     'AI talent-health matching + AI health tech skill assessment + AI interview prep chatbot + Health domain RAG',
     'Head of Health Vertical, CTO, Head of Talent', '~500', 'Series E'),
    ('Last Mile Health', 'Pan-Africa (Liberia, Ethiopia, etc.)',
     'Community health worker programs. Training & deploying health workers in remote areas',
     'CHWs need decision support in field. Training in low-resource settings. Patient data collection. Disease reporting',
     'AI CHW decision support chatbot (works offline) + AI training assistant + AI disease reporting + Health protocol RAG',
     'Head of Technology, CTO, Head of Programs', '~100', 'NGO'),
    ('Living Goods', 'Nairobi, Kenya / Uganda',
     'Community health platform. Tech-enabled community health workers serving 10M+ people',
     'CHWs need real-time guidance on diagnostics. Patient follow-up across villages. Stock management of health supplies',
     'AI diagnostic support chatbot + AI patient follow-up system + AI stock predictor + Community health RAG',
     'CTO, Head of Technology, Head of AI', '~200', 'NGO'),
    ('Dimagi (CommCare)', 'Pan-Africa',
     'Mobile health data platform. CommCare used by 100+ health programs across Africa',
     'Health programs need AI analytics. CommCare data underutilized. Program managers need AI insights. Report generation slow',
     'AI health program analytics + AI report generator + AI data quality checker + Program knowledge RAG chatbot',
     'CTO, Head of AI, Head of Product', '~200', 'Social Enterprise'),
    ('Praekelt (Turn.io)', 'Cape Town, South Africa',
     'Health messaging platform. WhatsApp-based health services. MomConnect (3M+ mothers in SA)',
     'Millions of health messages. Content personalization needed. Multi-language. Triage via chat. Health worker support',
     'AI multi-agent health messaging system + AI content personalization + AI triage chatbot + Multi-language NLP',
     'CTO, Head of AI, Head of Engineering', '~50', 'Social Enterprise'),
    ('Africa CDC (tech partners)', 'Pan-Africa',
     'Continental health body. Disease surveillance, pandemic response across 55 African countries',
     'Disease outbreak detection slow. Data from 55 countries. Reporting manual. Health worker coordination complex',
     'AI disease surveillance dashboard + AI outbreak prediction + AI reporting automation + Health data RAG',
     'Head of Digital, CTO, Chief Data Officer', '~200', 'Govt/Intl'),
    ('WHO AFRO Digital Health', 'Pan-Africa',
     'WHO Africa region digital health initiatives. Supporting 47 countries with health tech',
     'Health data standardization across countries. AI tools for health workers. Digital health training programs',
     'AI health data standardizer + AI training chatbot + AI program monitoring + Health guidelines RAG',
     'Digital Health Lead, Technology Advisor, CTO', '~500', 'Intl Org'),
    ('Partners In Health (PIH) - Africa', 'Rwanda, Sierra Leone, Liberia',
     'Healthcare delivery NGO. Clinics and hospitals in 3 African countries',
     'Clinical decision support for rural doctors. Patient records scattered. Drug supply management. Staff training',
     'AI clinical decision support + AI patient records RAG + AI drug supply tracker + AI staff training chatbot',
     'Head of Technology, CTO, Chief Medical Officer', '~200', 'NGO'),
]

assert len(companies) == 80, f"Got {len(companies)}"

NC = 17
ws1 = wb.active
ws1.title = 'Healthcare AI Tracker'
ws1.merge_cells(f'A1:{get_column_letter(NC)}1')
ws1.cell(1, 1, 'HEALTHCARE AI \u2014 8 COMPANIES/DAY x 10 DAYS = 80 TARGETS \u2014 Hasnain ka #1 Domain').font = Font(bold=True, size=12, color='1B3A5C')
ws1.row_dimensions[1].height = 28
ws1.merge_cells(f'A2:{get_column_letter(NC)}2')
ws1.cell(2, 1, 'Proof: CareBot (17 tables, 40+ endpoints, 16 AI tools) + HepaBot (Clinical BERT) + CL4 (FHIR) + PrimePal (3-agent chatbot)').font = Font(italic=True, color='548235', size=10)

hdrs = ['#', 'Day', 'Company', 'City', 'Kya Karti Hai',
        'UNKA DARD', 'KYA PITCH KARO',
        'Kon Dhundna Hai', 'Size', 'Stage',
        'Person Name', 'Role', 'LinkedIn URL',
        'Req?', 'Conn?', 'Msg?', 'Status']
for c, h in enumerate(hdrs, 1):
    ws1.cell(4, c, h)
style_hdr(ws1, 4, NC)
ws1.cell(4, 6).fill = PatternFill('solid', fgColor='BF4B28')
ws1.cell(4, 7).fill = pfill
for c in range(11, 18): ws1.cell(4, c).fill = pfill

row = 5; pnum = 0
for i, comp in enumerate(companies):
    dn = (i // 8) + 1
    dt = start + timedelta(days=dn - 1)
    if i % 8 == 0:
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NC)
        cl = ws1.cell(row, 1, f'DAY {dn} \u2014 {dt.strftime("%A %b %d")}')
        cl.font = dfont; cl.fill = dfill
        cl.alignment = Alignment(horizontal='center', vertical='center')
        ws1.row_dimensions[row].height = 25; row += 1
    name, city, does, pain, pitch, roles, size, stage = comp
    for j in range(4):
        pnum += 1
        ws1.cell(row, 1, pnum); ws1.cell(row, 2, f'Day {dn}'); ws1.cell(row, 3, name)
        if j == 0:
            ws1.cell(row, 4, city); ws1.cell(row, 5, does)
            ws1.cell(row, 6, pain); ws1.cell(row, 7, pitch)
            ws1.cell(row, 8, roles); ws1.cell(row, 9, size); ws1.cell(row, 10, stage)
        ws1.cell(row, 14, 'No'); ws1.cell(row, 15, 'No'); ws1.cell(row, 16, 'No')
        ws1.cell(row, 17, 'Not Started')
        f = lgray if j % 2 == 0 else None
        style_row(ws1, row, NC, f)
        if j == 0:
            ws1.row_dimensions[row].height = 75
            for c in range(3, 11): ws1.cell(row, c).font = Font(bold=True)
            ws1.cell(row, 6).font = Font(bold=True, color='BF4B28')
            ws1.cell(row, 7).font = Font(bold=True, color='548235')
        else:
            ws1.row_dimensions[row].height = 25
        row += 1
    for c in range(1, NC+1): ws1.cell(row-1, c).border = bb

dv_yn = DataValidation(type='list', formula1='"Yes,No"')
dv_st = DataValidation(type='list', formula1='"Not Started,Request Sent,Connected,Message Sent,Follow-Up,Replied,Call Booked,Converted,Dead"')
ws1.add_data_validation(dv_yn); ws1.add_data_validation(dv_st)
for r in range(5, row):
    v = ws1.cell(r, 1).value
    if v and isinstance(v, int):
        dv_yn.add(ws1.cell(r, 14)); dv_yn.add(ws1.cell(r, 15)); dv_yn.add(ws1.cell(r, 16))
        dv_st.add(ws1.cell(r, 17))

widths = [5, 7, 18, 16, 38, 42, 42, 30, 7, 8, 20, 18, 40, 6, 6, 6, 12]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A5'

# =========================================================
# SHEET 2: DOMAIN PRIORITY & ROADMAP
# =========================================================
ws2 = wb.create_sheet('Domain Priority & Roadmap')
ws2.merge_cells('A1:H1')
ws2.cell(1, 1, 'DOMAIN ROADMAP \u2014 HASNAIN KI STRENGTHS KE MUTABIQ PRIORITY').font = Font(bold=True, size=13, color='1B3A5C')
ws2.merge_cells('A2:H2')
ws2.cell(2, 1, 'Based on: Resume (Turing, i2c, nutsandbolts) + Thesis (PrimePal) + Portfolio (CareBot, DocMind, FinancePal, WhatsBot, LeadFlow)').font = Font(italic=True, color='666666', size=10)

# Section: CURRENT
r = 4
ws2.merge_cells(f'A{r}:H{r}')
ws2.cell(r, 1, 'ABHI KAR RAHE HO (CURRENT)').font = Font(bold=True, size=12, color='FFFFFF')
for c in range(1, 9): ws2.cell(r, c).fill = dfill; ws2.cell(r, c).border = tb
r += 1

ch = ['#', 'Domain', 'Status', 'Duration', 'Hasnain Ka Proof', 'Kya Pitch', 'Target Companies', 'Kyun Payenge']
for c, h in enumerate(ch, 1): ws2.cell(r, c, h)
style_hdr(ws2, r, 8); r += 1

ws2.cell(r, 1, '1'); ws2.cell(r, 2, 'HEALTHCARE AI')
ws2.cell(r, 3, 'IN PROGRESS'); ws2.cell(r, 4, '10 days (80 companies)')
ws2.cell(r, 5, 'CareBot (17 tables, 40+ endpoints, 16 AI tools, HIPAA)\nHepaBot (Clinical BERT + LoRA)\nCL4 (FHIR-native medical records)\nPrimePal (3-agent conversational AI)')
ws2.cell(r, 6, 'AI patient triage chatbot\nMedical knowledge RAG\nClinic management AI\nChronic disease chatbot')
ws2.cell(r, 7, '80 companies in Sheet 1')
ws2.cell(r, 8, 'VC-funded ($10M+). Going digital fast. HIPAA/FHIR knowledge RARE in Africa. Instant trust.')
style_row(ws2, r, 8, gfill); ws2.row_dimensions[r].height = 100; r += 2

# Section: NEXT UP
ws2.merge_cells(f'A{r}:H{r}')
ws2.cell(r, 1, 'NEXT UP \u2014 Healthcare ke baad yahan jaana hai').font = Font(bold=True, size=12, color='FFFFFF')
for c in range(1, 9): ws2.cell(r, c).fill = PatternFill('solid', fgColor='548235'); ws2.cell(r, c).border = tb
r += 1

for c, h in enumerate(ch, 1): ws2.cell(r, c, h)
style_hdr(ws2, r, 8); r += 1

next_domains = [
    ('2', 'EDTECH / E-LEARNING', 'NEXT', '10 days',
     'PrimePal = DEPLOYED, field-tested, thesis video recorded\n3-agent system with RAG, tested on real students\nNO other freelancer has this',
     'AI adaptive tutoring\nAI assessment engine\nAI student analytics\nAI pronunciation scoring',
     'uLesson, AltSchool Africa, SomaAI, Andela Learning, Decagon',
     'PrimePal story sells to startups AND NGOs/grants (GIZ, USAID). English tutoring = universal need in Africa'),
    ('3', 'RAG / DOCUMENT Q&A', 'NEXT', '5 days',
     'DocMind RAG = LIVE (ChromaDB, cited answers)\nPrimePal = pgvector RAG with guardrails\nWorks in EVERY industry',
     'AI document Q&A system\nKnowledge base chatbot\nCited answers with sources\nSmart doc search',
     'ANY startup with docs: Paystack, Flutterwave, SeamlessHR, law firms, hospitals, banks',
     'Saves 10+ hrs/week. ROI immediate. Every CTO understands "searching for answers" problem'),
    ('4', 'MCP / CLAUDE API', 'PLANNED', '5 days',
     'FinancePal = 14-tool MCP server, 22 tests\nSABSE KAM competition (5-20 proposals vs 50+)\nMost freelancers don\'t know MCP',
     'Custom MCP server development\nClaude API integration\nMulti-tool AI agents',
     'Any company adopting Claude, consulting firms, enterprise AI teams',
     'New tech = premium rates ($35-100/hr). First-mover advantage. Growing fast'),
    ('5', 'n8n AUTOMATION + AI', 'PLANNED', '5 days',
     'n8n AI Workflows = 3 production workflows DONE\nLeadFlow = CRM + AI lead scoring DONE\nBundles AI INTO workflow (others just plumbing)',
     'Email triage + auto-response\nInvoice OCR + extraction\nLead scoring + CRM sync',
     'Any SME: Brass, SeamlessHR, Termii, sales teams, finance teams',
     'Quick ROI ("saved 20 hrs/week"). Small ticket ($500-2K) but fast close. Door opener for bigger projects'),
    ('6', 'AI SaaS MVP BUILDING', 'PLANNED', '10 days',
     'PrimePal = full production system\nDual auth, multi-role dashboard, Redis, Docker, 123+ tests\n35+ DB migrations, field-deployed',
     'Build complete AI MVP in 4-8 weeks\nMulti-agent + RAG + chatbot + dashboard',
     'YC Africa alumni, TechStars, Founders Factory portfolio, angel-funded startups',
     '$5K-25K per MVP. Founders raised money TO BUILD. Timeline pressure = pay fast. Show PrimePal = instant credibility'),
]

for nd in next_domains:
    for c, v in enumerate(nd, 1): ws2.cell(r, c, v)
    style_row(ws2, r, 8, yfill); ws2.row_dimensions[r].height = 100; r += 1

r += 1
# Section: LATER
ws2.merge_cells(f'A{r}:H{r}')
ws2.cell(r, 1, 'LATER \u2014 Upar wale domains ke baad').font = Font(bold=True, size=12, color='FFFFFF')
for c in range(1, 9): ws2.cell(r, c).fill = PatternFill('solid', fgColor='808080'); ws2.cell(r, c).border = tb
r += 1

for c, h in enumerate(ch, 1): ws2.cell(r, c, h)
style_hdr(ws2, r, 8); r += 1

later = [
    ('7', 'Legal Tech', 'LATER', '5 days', 'DocMind RAG = legal doc search. High value clients', 'AI contract analyzer + Legal Q&A chatbot', 'LawPavilion, CaseTrack, Lex Artifex', 'Premium rates but lawyers conservative. Slow close'),
    ('8', 'E-Commerce / Retail', 'LATER', '5 days', 'RAG + WhatsBot = shopping assistant', 'AI product search + Recommendations + Chatbot', 'Jumia, Konga, Sabi, TradeDepot, Omnibiz', 'Big market but thin margins. Better for product features'),
    ('9', 'Agritech', 'LATER', '5 days', 'WhatsBot + RAG = farm advisory bot', 'AI farmer chatbot + Crop RAG + Market matcher', 'Releaf, ThriveAgric, Apollo Agri, Farmerline', 'Mission-aligned. VC-funded ones pay. Grant-dependent ones slow'),
    ('10', 'Insurtech', 'LATER', '5 days', 'Document RAG + chatbot = claims processing', 'AI claims bot + Policy explainer + Fraud detection', 'Turaco, Lami, Curacel, Casava, Hygeia', 'High-margin industry but African insurtech is early'),
]
for lt in later:
    for c, v in enumerate(lt, 1): ws2.cell(r, c, v)
    style_row(ws2, r, 8, lgray); ws2.row_dimensions[r].height = 60; r += 1

r += 1
ws2.merge_cells(f'A{r}:H{r}')
ws2.cell(r, 1, 'TOTAL ROADMAP: Healthcare (10 days) \u2192 Edtech (10) \u2192 RAG (5) \u2192 MCP (5) \u2192 n8n (5) \u2192 MVP (10) \u2192 Later (20) = ~65 days').font = Font(bold=True, size=11, color='1B3A5C')

for i, w in enumerate([4, 22, 12, 16, 50, 35, 45, 45], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

fname = 'Hasnain_Lagos_AI_SaaS_Outreach.xlsx'
wb.save(fname)
print(f'DONE: {fname}')
print(f'Sheet 1: Healthcare AI - {len(companies)} companies, {pnum} people rows')
print(f'Sheet 2: Domain Priority & Roadmap')
